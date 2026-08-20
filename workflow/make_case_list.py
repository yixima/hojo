#!/usr/bin/env python3
"""案件リストの CSV から配布用 Excel を作る。

    python3 workflow/make_case_list.py data/cases/cases_20260819.csv

出力:
    data/cases/cases_YYYYMMDD.xlsx   配布用（Mac / Windows の Excel で文字化けしない）
    data/cases/cases_YYYYMMDD.csv    BOM付きUTF-8に書き直す（Excelで直接開ける）

CSV を UTF-8（BOMなし）のまま Mac の Excel で開くと文字化けする。
Excel は BOM が無い CSV を実行環境の既定エンコーディングとみなすため。
配布は Excel を主とし、CSV は BOM 付きで併置する。

残日数と緊急度は Excel の数式にしてある。H1 の「基準日」を今日の日付に
書き換えれば全行が再計算される。ハードコードしないこと。
"""
import csv
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Yu Gothic"          # 和文を含むため。Arial だと和文がフォールバックして不揃いになる

INK        = "14202E"
INDIGO     = "1E4A7B"
INDIGO_DP  = "14304F"
VERMILLION = "BE3A2B"
MUSTARD    = "8E6A16"
CELADON    = "2C6D5E"
GREY       = "6B7887"
TINT_V     = "FAE6E3"
TINT_M     = "F5EEDA"
TINT_C     = "DFEDE8"
TINT_I     = "E2EAF3"
BAND       = "F5F7FA"
YELLOW     = "FFFF00"

# (見出し, 幅, 折返し)
COLUMNS = [
    ("No.",                 5,  False),
    ("区分",                 8,  False),
    ("格付",                 6,  False),
    ("緊急度",               9,  False),
    ("残日数",               8,  False),
    ("案件名",               46, True),
    ("種別",                 13, False),
    ("発注機関",             24, True),
    ("見込み金額",           15, False),
    ("金額確度",             18, False),
    ("金額の根拠",           46, True),
    ("公告日",               12, False),
    ("締切",                 17, False),
    ("直近アクション期限",    17, False),
    ("直近アクションの内容",  34, True),
    ("推奨着手日",           12, False),
    ("予測公告時期",         16, True),
    ("予測の根拠",           40, True),
    ("応募形態",             12, False),
    ("資格要否",             26, True),
    ("前提条件",             34, True),
    ("実績要件",             34, True),
    ("判定",                 10, False),
    ("次アクション",         34, True),
    ("状態",                 10, False),
    ("レポート",             12, False),
    ("一次情報URL",          40, False),
]

RANK_COLOR   = {"S": VERMILLION, "A": INDIGO, "B": INK, "C": GREY}
VERDICT_COLOR = {"応募推奨": CELADON, "要検討": INDIGO, "記録のみ": GREY, "見送り": GREY}
URGENCY_FILL = {"至急": TINT_V, "要着手": TINT_M, "通常": TINT_C}
URGENCY_TEXT = {"至急": VERMILLION, "要着手": MUSTARD, "通常": CELADON}

thin = Side(style="thin", color="D3D9E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def parse_dt(s):
    """'2026-08-25 09:00' / '2026-08-25' / '' を datetime|date|None へ。"""
    s = (s or "").strip()
    if not s:
        return None
    for fmt, is_date in (("%Y-%m-%d %H:%M", False), ("%Y-%m-%d", True)):
        try:
            d = datetime.strptime(s, fmt)
            return d.date() if is_date else d
        except ValueError:
            continue
    return None


def build(csv_path: Path) -> Path:
    rows = list(csv.DictReader(csv_path.open(encoding="utf-8-sig")))
    stamp = csv_path.stem.replace("cases_", "")
    base_date = datetime.strptime(stamp, "%Y%m%d").date()

    wb = Workbook()
    # この環境の LibreOffice は動作しないため recalc.py で値を焼き込めない。
    # openpyxl は数式のキャッシュ値を書けないので、Excel 側に開いた時点での
    # 全再計算を強制する。これが無いと一部のビューアで空欄に見える。
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "案件リスト"

    # ---- タイトル行 -------------------------------------------------
    ws["A1"] = f"公募案件リスト {base_date:%Y.%m.%d}"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INDIGO_DP)
    ws.merge_cells("A1:E1")

    ws["G1"] = "基準日 →"
    ws["G1"].font = Font(name=FONT, size=10, bold=True, color=INK)
    ws["G1"].alignment = Alignment(horizontal="right", vertical="center")
    ws["H1"] = base_date
    ws["H1"].number_format = "yyyy-mm-dd"
    ws["H1"].font = Font(name=FONT, size=10, bold=True, color=INK)
    ws["H1"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["H1"].border = BORDER
    ws["I1"] = "← ここを今日の日付に変えると 残日数 と 緊急度 が再計算されます"
    ws["I1"].font = Font(name=FONT, size=9, color=GREY)
    ws.row_dimensions[1].height = 22

    # ---- 見出し行（3行目） -------------------------------------------
    HEAD = 3
    for i, (name, width, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HEAD, column=i, value=name)
        c.font = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INDIGO_DP)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[HEAD].height = 30

    # ---- データ行 ----------------------------------------------------
    for n, r in enumerate(rows):
        y = HEAD + 1 + n
        rank = r["格付"].strip()
        verdict = r["判定"].strip()
        amount = r["見込み金額"].strip()
        band = BAND if n % 2 else None

        def put(col, value, *, fmt=None, color=INK, bold=False, wrap=False, size=9.5):
            c = ws.cell(row=y, column=col, value=value)
            c.font = Font(name=FONT, size=size, bold=bold, color=color)
            c.alignment = Alignment(vertical="top", wrap_text=wrap, horizontal="left")
            c.border = BORDER
            if fmt:
                c.number_format = fmt
            if band:
                c.fill = PatternFill("solid", fgColor=band)
            return c

        kind = r.get("区分", "公募中").strip()
        put(1, int(r["No."] if "No." in r else r["No"]))
        put(2, kind, color=INDIGO if kind == "公募中" else MUSTARD, bold=True)
        put(3, rank, color=RANK_COLOR.get(rank, INK), bold=True)

        # 緊急度・残日数は数式。基準日(H1)を変えれば追随する。
        # 予測案件は期限が無いので空欄になる。
        put(4, f'=IF(E{y}="","",IF(E{y}<=7,"至急",IF(E{y}<=21,"要着手","通常")))', bold=True)
        put(5, f'=IF(N{y}="","",INT(N{y})-INT($H$1))', fmt="0")

        put(6, r["案件名"], wrap=True)
        put(7, r["種別"], color=GREY)
        put(8, r["発注機関"], wrap=True)

        if amount:
            put(9, int(amount), fmt='¥#,##0;;"—"', bold=True)
        else:
            put(9, "—", color=GREY)
        conf = r["金額確度"]
        put(10, conf, color=CELADON if conf.startswith("◎") else MUSTARD if conf.startswith("△") else GREY)

        put(11, r.get("金額の根拠", ""), wrap=True, size=8.5, color=GREY)
        put(12, parse_dt(r["公告日"]), fmt="yyyy-mm-dd", color=GREY)
        put(13, parse_dt(r["締切"]), fmt="yyyy-mm-dd hh:mm", bold=True)
        put(14, parse_dt(r["直近アクション期限"]), fmt="yyyy-mm-dd hh:mm", bold=True)
        put(15, r["直近アクションの内容"], wrap=True, size=9)
        put(16, parse_dt(r.get("推奨着手日", "")), fmt="yyyy-mm-dd", color=MUSTARD)
        put(17, r.get("予測公告時期", ""), wrap=True, color=MUSTARD, bold=True)
        put(18, r.get("予測の根拠", ""), wrap=True, size=9, color=GREY)

        put(19, r["応募形態"])
        q = r["資格要否"]
        put(20, q, wrap=True, size=9,
            color=VERMILLION if "間に合わない" in q else MUSTARD if "代替可" in q else GREY)
        put(21, r["前提条件"], wrap=True, size=9)
        put(22, r["実績要件"], wrap=True, size=9)
        put(23, verdict, color=VERDICT_COLOR.get(verdict, INK), bold=True)
        put(24, r["次アクション"], wrap=True, size=9)
        put(25, r["状態"])

        link = r["レポート節"]
        c = put(26, "該当節へ", color="0563C1")
        if link:
            c.hyperlink = REPORT_URL + link
            c.font = Font(name=FONT, size=9.5, color="0563C1", underline="single")

        url = r["一次情報URL"]
        c = put(27, url, color="0563C1", size=8.5)
        if url:
            c.hyperlink = url
            c.font = Font(name=FONT, size=8.5, color="0563C1", underline="single")

        ws.row_dimensions[y].height = 44

    last = HEAD + len(rows)
    ws.auto_filter.ref = f"A{HEAD}:AA{last}"
    ws.freeze_panes = f"G{HEAD + 1}"

    dv = DataValidation(type="list",
                        formula1='"新規,検討中,提出済,見送り,締切超過"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"Y{HEAD + 1}:Y{last}")

    # ---- 凡例シート --------------------------------------------------
    write_legend(wb, base_date)

    out = csv_path.with_suffix(".xlsx")
    wb.save(out)
    return out


REPORT_URL = "https://claude.ai/code/artifact/6f25d20f-a5ce-4db2-b4b2-5e46415b52a6"

LEGEND = [
    ("No.", "その回の中での通し番号"),
    ("区分", "公募中 = いま応募できるもの / 予測 = 過去の公告実績から近く出ると見込まれるもの"),
    ("格付", "S / A / B / C。適合度。workflow/eligibility.md [4] の基準"),
    ("緊急度", "【数式】残日数から自動判定。7日以下=至急 / 8〜21日=要着手 / 22日以上=通常"),
    ("残日数", "【数式】直近アクション期限 − 基準日(H1)。締切までの日数ではない"),
    ("案件名", "公告の表記をそのまま"),
    ("種別", "補助金 / 企画競争 / 企画提案 / プロポーザル / 入札 / 委託"),
    ("発注機関", "正式名称"),
    ("見込み金額", "補助金は補助上限額、それ以外は予定価格。不明は「—」"),
    ("金額確度", "◎確定 / △推定 / －非公開。空欄と「調べていない」を区別するため必須"),
    ("金額の根拠", "どこの何を見てその金額にしたか。非公開ならどこを確認して非公開と判断したか"),
    ("公告日", ""),
    ("締切", "時刻まで記載。17:00必着と14:00では動きが変わる"),
    ("直近アクション期限", "締切より前に来る公表期限（説明会申込・質問締切・資格申請など）。"
                          "無ければ締切と同値。緊急度はこの列で測る"),
    ("直近アクションの内容", "その期限までに何をするか"),
    ("推奨着手日", "こちらの判断で置いた目安。公表期限ではないため緊急度の計算には使わない"),
    ("予測公告時期", "区分=予測 のみ。監視を始めるべき月"),
    ("予測の根拠", "区分=予測 のみ。何年連続で・いつ公告されたか。根拠の無い予測は載せない"),
    ("応募形態", "単独 / 共同 / 支援側 / 受託"),
    ("資格要否", "不要 / 要（間に合う）/ 要（代替可）/ 要（間に合わない）。"
                 "「代替可」は等級確認等の迂回路がある場合"),
    ("前提条件", "GビズID、Pマーク・ISMS、ナビ登録、現地体制 など"),
    ("実績要件", "○ / △ / × と根拠。△は「要領未取得で未確認」を含む"),
    ("判定", "応募推奨 / 要検討 / 記録のみ / 見送り"),
    ("次アクション", "次に取る動作。無い場合は「対応不要」"),
    ("状態", "新規 / 検討中 / 提出済 / 見送り / 締切超過。data/ledger.csv と同じ語彙"),
    ("レポート", "レポート本文の該当節へのリンク"),
    ("一次情報URL", "発注元の原典。集約サイトのURLは不可"),
]


def write_legend(wb, base_date):
    ws = wb.create_sheet("凡例・項目定義")
    ws["A1"] = "項目定義"
    ws["A1"].font = Font(name=FONT, size=13, bold=True, color=INDIGO_DP)

    ws["A3"] = "編集してよいセル"
    ws["A3"].font = Font(name=FONT, size=10, bold=True, color=INK)
    notes = [
        "「案件リスト」シートの H1（黄色）… 基準日。今日の日付に変えると"
        "「残日数」と「緊急度」が全行再計算されます",
        "「案件リスト」シートの U列 … 状態。ドロップダウンから選べます",
        "それ以外の列は巡回結果です。上書きすると次回の突き合わせができなくなります",
    ]
    for i, t in enumerate(notes):
        c = ws.cell(row=4 + i, column=1, value=("・" + t))
        c.font = Font(name=FONT, size=9.5, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=4 + i, start_column=1, end_row=4 + i, end_column=2)
        ws.row_dimensions[4 + i].height = 28

    head = 4 + len(notes) + 1
    for i, name in enumerate(("列", "意味"), start=1):
        c = ws.cell(row=head, column=i, value=name)
        c.font = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INDIGO_DP)
        c.border = BORDER
    for i, (name, desc) in enumerate(LEGEND):
        y = head + 1 + i
        a = ws.cell(row=y, column=1, value=name)
        a.font = Font(name=FONT, size=9.5, bold=True, color=INK)
        a.alignment = Alignment(vertical="top")
        a.border = BORDER
        b = ws.cell(row=y, column=2, value=desc)
        b.font = Font(name=FONT, size=9.5, color=INK)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        b.border = BORDER
        ws.row_dimensions[y].height = 30 if len(desc) > 46 else 18
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78

    y = head + len(LEGEND) + 2
    ws.cell(row=y, column=1, value="出典").font = Font(name=FONT, size=9.5, bold=True, color=INK)
    for i, t in enumerate((
        f"巡回日 {base_date:%Y-%m-%d}。項目の仕様は workflow/case-list-schema.md",
        f"レポート本文 {REPORT_URL}",
        "金額・期限はすべて公募要領・公告の原文で確認。推定値は金額確度に「△推定」を入れています",
    )):
        c = ws.cell(row=y + 1 + i, column=1, value="・" + t)
        c.font = Font(name=FONT, size=9, color=GREY)
        ws.merge_cells(start_row=y + 1 + i, start_column=1, end_row=y + 1 + i, end_column=2)


def rewrite_csv_with_bom(csv_path: Path):
    """Excel が直接開けるよう BOM 付き UTF-8 に書き直す。"""
    text = csv_path.read_text(encoding="utf-8-sig")
    csv_path.write_text(text, encoding="utf-8-sig")


def verify(csv_path: Path):
    """数式のロジックを Python 側で独立に再現し、CSV のスナップショット値と突き合わせる。

    LibreOffice が使えない環境では recalc.py で検算できないため、
    せめて数式の意味が正しいことをここで担保する。
    """
    rows = list(csv.DictReader(csv_path.open(encoding="utf-8-sig")))
    base = datetime.strptime(csv_path.stem.replace("cases_", ""), "%Y%m%d").date()
    bad = []
    for r in rows:
        if r.get("区分", "公募中").strip() == "予測":
            continue                      # 予測案件に期限は無い
        act = parse_dt(r["直近アクション期限"])
        if act is None:
            continue
        d = (act.date() if isinstance(act, datetime) else act) - base
        days = d.days
        urg = "至急" if days <= 7 else "要着手" if days <= 21 else "通常"
        if str(days) != r["残日数"].strip():
            bad.append(f'No.{r["No."]} 残日数 期待{days} / CSV {r["残日数"]}')
        if urg != r["緊急度"].strip():
            bad.append(f'No.{r["No."]} 緊急度 期待{urg} / CSV {r["緊急度"]}')
    if bad:
        raise SystemExit("数式と CSV が不一致:\n  " + "\n  ".join(bad))
    print(f"verified {len(rows)} rows: 残日数・緊急度の数式ロジックが CSV と一致")


if __name__ == "__main__":
    p = Path(sys.argv[1])
    rewrite_csv_with_bom(p)
    verify(p)
    out = build(p)
    print(f"wrote {out}")
    print(f"rewrote {p} as UTF-8 with BOM")
